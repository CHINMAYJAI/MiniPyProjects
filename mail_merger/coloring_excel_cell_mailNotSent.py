# NOTE: this script file is used to color the cells which shows that mail is not sent

# importing necessary files
import extract_not_sent_rows
# importing necessary modules
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def coloringCells():
    """Function uses to color the cells which are having the value NOT SENT in the column MailStatus"""
    # Load Excel file
    file_path = 'MailRecipients.xlsx'
    df = pd.read_excel(file_path)

    # Remove unnamed columns
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

    # Save cleaned DataFrame to a temporary Excel file
    temp_file = 'temp_mail.xlsx'
    df.to_excel(temp_file, index=False)

    # Load with openpyxl for formatting
    wb = load_workbook(temp_file)
    ws = wb.active

    # Define the red fill for cells with "not sent"
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Find the column index of 'MailStatus'
    column_name = 'MailStatus'
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' not found in Excel file.")

    col_idx = df.columns.get_loc(column_name) + 1  # Excel column index (1-based)

    # Loop through rows and apply formatting
    for row_num, value in enumerate(df[column_name], start=2):  # start=2 accounts for header row
        if str(value).strip().lower() == 'not sent':
            ws.cell(row=row_num, column=col_idx).fill = red_fill

    # Save formatted Excel back to original file
    wb.save(file_path)

    # Delete temporary file
    if os.path.exists(temp_file):
        os.remove(temp_file)
    
    # Calling extract_not_sent_rows file
    extract_not_sent_rows.filteringMailStatusAndCreationOfNewExcelFile()
if __name__ == "__main__":
    pass